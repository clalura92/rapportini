import os
import re
import json
import shutil
import requests
import openpyxl
import fitz
import spire.xls as spire
from spire.xls import FileFormat

_ON_RENDER = os.environ.get('RENDER') is not None
OUTPUT_PEVE   = '/tmp/Output_Rapportini_Peve/' if _ON_RENDER else 'Output_Rapportini_Peve/'
OUTPUT_FAUSTO = '/tmp/Output_Rapportini_Fausto/' if _ON_RENDER else 'Output_Rapportini_Fausto/'

_SPIRE_WARN_RE = re.compile(
    rb'BT /FAAA\w+ \S+ Tf \S+ \S+ \S+ \S+ \S+ \S+ Tm 1 0 0 rg \[.+?\] TJ\s+ET',
    re.DOTALL,
)

_LOGO_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'templates', 'logo_solware.png')
_LOGO_RECT = fitz.Rect(620, 77, 673, 117)

_XLSX_CONTEXT_LIMIT = 6000


def resolve_paths(root, report_type, year, month, task_category, partner_name, project_name):
    _proj = project_name if project_name else ''
    _sep  = ' - ' if _proj else ''
    basename = (f'{year}-{month} _' + task_category.upper() + '_'
                + re.split(r'[ ]', partner_name)[0] + _sep + _proj)

    if report_type == 'peve':
        out_dir = os.path.join(root, OUTPUT_PEVE, f'{year}_{month}')
    else:
        out_dir = os.path.join(root, OUTPUT_FAUSTO, f'{year}_{month}')

    xlsx_path = os.path.join(out_dir, basename + '.xlsx')
    pdf_path  = os.path.join(out_dir, basename + '.pdf')
    return xlsx_path, pdf_path


def read_xlsx_as_context(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    lines = []
    for sheet in wb.worksheets:
        lines.append(f'Sheet: {sheet.title}')
        for row in sheet.iter_rows():
            row_parts = []
            for cell in row:
                if cell.value is not None and str(cell.value).strip():
                    row_parts.append(f'{cell.coordinate}: {cell.value}')
            if row_parts:
                lines.append('  ' + '   '.join(row_parts))

    context = '\n'.join(lines)
    if len(context) > _XLSX_CONTEXT_LIMIT:
        context = context[:_XLSX_CONTEXT_LIMIT] + '\n[... truncated for context window]'
    return context


def call_gemini(context, history, user_message):
    api_key = os.environ.get('GEMINI_API_KEY', '')
    if not api_key:
        raise ValueError('GEMINI_API_KEY non configurato')

    system_prompt = f"""Sei un assistente che aiuta a modificare file Excel di rapportini di lavoro.

--- CONTENUTO EXCEL ATTUALE ---
{context}
--- FINE CONTENUTO EXCEL ---

Regole:
1. Se l'utente fa una domanda, rispondi direttamente. Imposta action su "answer".
2. Se l'utente chiede di modificare qualcosa, restituisci le modifiche. Imposta action su "modify".
3. Se l'utente chiede di annullare, ripristinare, tornare indietro, revertire o fare "undo" dell'ultima modifica, imposta action su "revert".
4. Rispondi SEMPRE con JSON valido in esattamente questa struttura:
{{
  "action": "answer" oppure "modify" oppure "revert",
  "message": "risposta in testo semplice per l'utente",
  "changes": [{{"sheet": 0, "cell": "B13", "value": "nuovo valore"}}]
}}
"changes" deve essere una lista vuota quando action è "answer" o "revert". L'indice dello sheet è 0-based. La notazione delle celle è standard Excel (es. "B13")."""

    contents = []
    for msg in history:
        role = 'model' if msg.get('role') == 'assistant' else 'user'
        # For model turns use the raw JSON Gemini actually produced so context is exact.
        # Fall back to human text if raw_response is not available (older messages).
        text = msg.get('raw_response') or msg.get('text', '')
        contents.append({'role': role, 'parts': [{'text': text}]})

    contents.append({'role': 'user', 'parts': [{'text': user_message}]})

    url = f'https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash:generateContent?key={api_key}'
    resp = requests.post(url, json={
        'systemInstruction': {'parts': [{'text': system_prompt}]},
        'contents': contents,
    }, timeout=60)
    resp.raise_for_status()

    raw_text = resp.json()['candidates'][0]['content']['parts'][0]['text']
    raw_text = raw_text.strip()
    if raw_text.startswith('```'):
        raw_text = re.sub(r'^```[a-z]*\n?', '', raw_text)
        raw_text = re.sub(r'\n?```$', '', raw_text)

    try:
        result = json.loads(raw_text)
    except json.JSONDecodeError:
        result = {'action': 'answer', 'message': raw_text, 'changes': []}
        raw_text = json.dumps(result, ensure_ascii=False)

    return result, raw_text


def _backup_paths(xlsx_path, pdf_path):
    return xlsx_path + '.bak', pdf_path + '.bak'


def create_backup(xlsx_path, pdf_path):
    from storage import upload_file, object_exists, to_storage_key
    bak_xlsx_key = to_storage_key(xlsx_path) + '.bak'
    bak_pdf_key  = to_storage_key(pdf_path)  + '.bak'
    if not object_exists(bak_xlsx_key):
        upload_file(xlsx_path, bak_xlsx_key)
    if os.path.exists(pdf_path) and not object_exists(bak_pdf_key):
        upload_file(pdf_path, bak_pdf_key)


def restore_backup(xlsx_path, pdf_path):
    from storage import download_to_bytes, object_exists, delete_file, to_storage_key
    bak_xlsx_key = to_storage_key(xlsx_path) + '.bak'
    bak_pdf_key  = to_storage_key(pdf_path)  + '.bak'
    if not object_exists(bak_xlsx_key):
        raise FileNotFoundError('Nessun backup disponibile per questo file')
    os.makedirs(os.path.dirname(xlsx_path), exist_ok=True)
    with open(xlsx_path, 'wb') as f:
        f.write(download_to_bytes(bak_xlsx_key))
    if object_exists(bak_pdf_key):
        with open(pdf_path, 'wb') as f:
            f.write(download_to_bytes(bak_pdf_key))
    delete_file(bak_xlsx_key)
    if object_exists(bak_pdf_key):
        delete_file(bak_pdf_key)


def apply_changes(xlsx_path, pdf_path, changes):
    create_backup(xlsx_path, pdf_path)
    wb = openpyxl.load_workbook(xlsx_path)
    for change in changes:
        sheet_idx = int(change.get('sheet', 0))
        cell_addr = change.get('cell', '')
        value     = change.get('value')
        if not cell_addr:
            continue
        if sheet_idx >= len(wb.worksheets):
            raise ValueError(f'Sheet index {sheet_idx} fuori range (il file ha {len(wb.worksheets)} sheet)')
        wb.worksheets[sheet_idx][cell_addr].value = value
    wb.save(xlsx_path)


def _clean_spire_pdf(pdf_path):
    tmp_path = pdf_path + '.tmp'
    doc = fitz.open(pdf_path)
    if doc.page_count > 1:
        doc.delete_page(-1)
    for page in doc:
        img_names = {img[7].encode() for img in page.get_images(full=True)}
        for xref in page.get_contents():
            raw = doc.xref_stream(xref)
            modified = raw
            if b'1 0 0 rg' in modified:
                modified = _SPIRE_WARN_RE.sub(b'', modified)
            for name in img_names:
                modified = re.sub(rb'/' + re.escape(name) + rb'\s+Do\b', b'', modified)
            if modified != raw:
                doc.update_stream(xref, modified)
        if os.path.exists(_LOGO_PATH):
            page.insert_image(_LOGO_RECT, filename=_LOGO_PATH, keep_proportion=True)
    doc.save(tmp_path, garbage=4, deflate=True)
    doc.close()
    os.replace(tmp_path, pdf_path)


def regenerate_pdf(xlsx_path, pdf_path):
    wb = spire.Workbook()
    wb.LoadFromFile(xlsx_path)
    wb.SaveToFile(pdf_path, FileFormat.PDF)
    wb.Dispose()
    _clean_spire_pdf(pdf_path)
    try:
        from storage import upload_file, to_storage_key
        upload_file(xlsx_path, to_storage_key(xlsx_path))
        upload_file(pdf_path,  to_storage_key(pdf_path))
    except Exception as e:
        print(f'Supabase upload failed (non-fatal): {e}')
