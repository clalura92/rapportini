import numpy as np
from math import ceil
import pandas as pd
import re
from datetime import datetime
from datetime import timedelta
import calendar as cal

import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.cell import MergedCell
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Color, PatternFill, Alignment
from openpyxl.worksheet.pagebreak import Break
from openpyxl.drawing.image import Image
from openpyxl.styles.borders import Border, Side
from openpyxl.drawing.spreadsheet_drawing import AbsoluteAnchor
from openpyxl.drawing.xdr import XDRPoint2D, XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU, cm_to_EMU

# NOTE: Spire.XLS (the embedded .NET CLR, ~100MB) and PyMuPDF (fitz) are imported
# lazily inside the functions that need them (create_rapportini / _clean_spire_pdf)
# rather than at module top. Listing projects imports this module but only uses the
# pure-pandas functions (load_df / list_projects / filter_*), so keeping these heavy
# imports out of module scope means a `projects` job never loads the .NET runtime.
import re
import os
from os import listdir
from os.path import isfile, join
import string
import shutil
import glob
import gc

import memlog


def is_daylight(date):
    # Last Sunday of March and last Sunday of October (EU DST rules)
    def last_sunday(y, m):
        last_day = cal.monthrange(y, m)[1]
        d = datetime(y, m, last_day)
        return d - timedelta(days=(d.weekday() + 1) % 7)
    start = last_sunday(date.year, 3)
    end = last_sunday(date.year, 10)
    return 2 if start <= date <= end else 1


def load_df(path_output, year, month):
    print(f"--- Starting load_df for {year}-{month} ---")
    file_path = path_output + str(year) + '_' + str(month) + '_timesheets_extraction.csv' 
    
    try:
        df = pd.read_csv(file_path)
        # Fix for Unix: strip any hidden whitespace from column names
        df.columns = df.columns.str.strip()
        print(f"DEBUG: Loaded {len(df)} rows.")
    except Exception as e:
        print(f"!!! Error reading CSV: {e}")
        return None

    # 1. Date and Time Processing (Do this BEFORE filtering)
    df['date'] = pd.to_datetime(df['date'], format='%Y-%m-%d')
    df['year'] = df['date'].dt.year
    df['month'] = df['date'].dt.month
    
    # We must ensure date_start/stop are datetime objects to use .dt.strftime
    df['time_diff'] = df['date'].apply(lambda x: is_daylight(x))
    
    print("DEBUG: Converting date_start/stop strings to datetime...")
    df['date_start'] = pd.to_datetime(df['date_start']) + pd.to_timedelta(df['time_diff'], unit='h')
    df['date_stop'] = pd.to_datetime(df['date_stop']) + pd.to_timedelta(df['time_diff'], unit='h')

    # 2. Create the columns that you use for sorting LATER
    df['hour_start'] = df['date_start'].dt.strftime('%H:%M')
    df['hour_stop'] = df['date_stop'].dt.strftime('%H:%M')
    df['time_period'] = df['hour_start'] + '-' + df['hour_stop']

    # 3. Safe Regex Extraction
    def get_match(val, pattern):
        match = re.search(pattern, str(val))
        return match.group(1) if match else "Unknown"

    print("DEBUG: Extracting names via Regex...")
    df['employee_name'] = df['employee_id'].apply(lambda x: get_match(x, ", '(.*) - "))
    df['partner_name'] = df['partner_id'].apply(lambda x: string.capwords(get_match(x.replace('"', "'"), ", '(.*)']")))
    df['project_name_orig'] = df['project_id'].apply(lambda x: get_match(x, r" - (.*) \[",))

    # 4. Filtering (Now safe because hour_start already exists)
    print("DEBUG: Filtering tasks...")
    valid_tasks = ['Intervento', 'Viaggio', 'Assistenza']
    df = df[df['task_name'].isin(valid_tasks)].copy()
    
    if df.empty:
        print(f"!!! WARNING: No rows match tasks {valid_tasks}. Check for case-sensitivity.")
        return df

    df['task_category'] = df['task_name'].apply(lambda x: 'Assistenza' if x=='Assistenza' else 'Intervento')
    
    # 5. Final filter by Year/Month and Sort
    df = df[(df['year'] == int(year)) & (df['month'] == int(month))]
    
    print(f"DEBUG: Rows remaining after year/month filter: {len(df)}")
    
    if not df.empty:
        df = df.sort_values('hour_start')
    
    print("--- load_df finished ---")
    return df


def hours_by_task_name(row, task_name):
    if row['task_name'] in task_name:
        return row['duration_unit_amount']
    else:
        return None


def clear_cells(ws):
    for row in ws['B7:I36']:
        for cell in row:
            cell.value = None
    for row in ws['N7:N36']:
        for cell in row:
            cell.value = None

    merged_cells_ranges = ws.merged_cells.ranges
    for merged in merged_cells_ranges:
        ws.unmerge_cells(str(merged))
    for row in ws['J7:J36']:
        for cell in row:
            cell.value = None
    for i in range(6, 37):
        ws.merge_cells('J' + str(i) + ':L' + str(i))

    return ws


def weekday_or_weekend(row):
    weekno = row['date'].weekday()
    if weekno < 5:
        return ''
    else:
        return 'F'


def clean_description(row, field):
    name = row[field]
    name = name.replace('<p>.</p>', '')
    name = name.replace('</p>', '')
    name = name.replace('<p>', '')
    name = name.replace('&nbsp', '')
    name = name.replace('<br>', '')
    name = name.replace('<span style="width: 40px;" class="oe-tabs">\t</span>\u200b', '')
    name = name.replace('<span style="width: 40px;" class="oe-tabs"> </span>\u200b', '')
    name = name.replace('<span class="oe-tabs" style="width: 40px;">	</span>', '')
    return name


def filter_df(df, employees, partner_name, project_name, task_category, flag_to_isolate):
    df = df[df['employee_name'].isin(employees)]
    if flag_to_isolate:
        df = df[(df['partner_name'] == partner_name) & (df['project_name'] == project_name) & (df['task_category'] == task_category)]
    else:
        df = df[(df['partner_name'] == partner_name) & (df['task_category'] == task_category)]
    df = df.sort_values(['task_name'], ascending = False)
    return df


def next_row_num(j, num_sheet, row_starts, row_ends, prev_employee, employee, df_pages):
    print(' --- funct next_row_num started')
    if j in row_ends:
        print(' --- funct next_row_num - j in row_ends started')
        print('num_sheet: ', num_sheet)
        num_sheet = num_sheet+1
        print('num_sheet_after: ', num_sheet)
        j = row_starts[num_sheet]
        print('Row_start[num_sheet]: ', row_starts[num_sheet])
        print('num_sheet: ', num_sheet)
        
        df_page_new = pd.DataFrame({'page': [num_sheet], 'employee': [employee]})
        df_pages = pd.concat([df_pages, df_page_new], ignore_index=True)
        print(' --- funct next_row_num - j in row_ends finished')
    elif prev_employee == '':
        print(' --- funct next_row_num - prev_employee == started')
        j += 1
        
        print(f"{j}, {num_sheet}, {employee}, {df_pages}")
        df_page_new = pd.DataFrame({'page': [num_sheet], 'employee': [employee]})
        df_pages = pd.concat([df_pages, df_page_new], ignore_index=True)
        
        print(f"{j}, {num_sheet}, {employee}, {df_pages}")
        print(' --- funct next_row_num - prev_employee == finished')
    elif prev_employee != employee:
        print(' --- funct next_row_num - prev_employee != started')
        num_sheet = num_sheet+1
        j = row_starts[num_sheet]
        
        print(f"{j}, {num_sheet}, {employee}, {df_pages}")
        df_page_new = pd.DataFrame({'page': [num_sheet], 'employee': [employee]})
        df_pages = pd.concat([df_pages, df_page_new], ignore_index=True)
        
        print(' --- funct next_row_num - prev_employee != finished')
    else:
        print(' --- funct next_row_num - j += 1 started')
        j += 1
        print(' --- funct next_row_num - j += 1 finished')
    
    return j, num_sheet, df_pages


def create_folders(path, year, month, owner):
    path_output = path + year + '_' + month + '/'
    os.makedirs(path_output, exist_ok=True)


def filter_employee_and_partners(df, eligibility_rules, dict_partner_rename, to_isolate_list):
    df['project_name'] = df.apply(lambda x: x['project_name_orig'] if x['partner_name'] in to_isolate_list else '', axis=1)
    print('df pre employee filter: ',df.shape)
    
    eligible_employees = [k for (k, v) in eligibility_rules.items()]
    df = df[df['employee_name'].isin(eligible_employees)]
    print('df post employee filter: ',df.shape)
    
    df['eligible_partners'] = df['employee_name'].map(eligibility_rules)
    df['eligibility'] = df.apply(lambda x: x['eligible_partners']==['*'] or (x['partner_name'] in x['eligible_partners']), axis=1)
    df = df[df['eligibility']==True]
    print('df post eligibily filter: ',df.shape)
    
    df['partner_name'] = df['partner_name'].map(dict_partner_rename).fillna(df['partner_name'])
    print('df post partner mapping: ',df.shape)

    return df


def identify_suspicious_data(df):
    flag_suspicious = []
    daterange = []
    
    for index, row in df.iterrows():
        project = row['project_name']
        days_to_check = [row['date'] - timedelta(days=1), row['date'], row['date'] + timedelta(days=1)]
        flag_suspicious += [('Intervento' not in df[(df['date'].isin(days_to_check)) & (df['partner_name'] == row['partner_name']) & (df['employee_name'] == row['employee_name'])]['task_name'].unique()) & (row['task_name'] == 'Viaggio')] 
    df['flag_suspicious'] = flag_suspicious
    
    return df


_SPIRE_WARN_RE = re.compile(
    rb'BT /FAAA\w+ \S+ Tf \S+ \S+ \S+ \S+ \S+ \S+ Tm 1 0 0 rg \[.+?\] TJ\s+ET',
    re.DOTALL,
)

_LOGO_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'templates', 'logo_solware.png')
_LOGO_RECT_COORDS = (620, 77, 673, 117)  # built into a fitz.Rect lazily in _clean_spire_pdf

def _clean_spire_pdf(pdf_path):
    import fitz  # PyMuPDF — heavy; imported here so listing never loads it
    tmp_path = pdf_path + '.tmp'
    doc = fitz.open(pdf_path)
    if doc.page_count > 1:
        doc.delete_page(-1)
    for page in doc:
        img_names = {img[7].encode() for img in page.get_images(full=True)}
        for xref in page.get_contents():
            raw = doc.xref_stream(xref)
            modified = raw
            if b'1 0 0 rg' in modified:
                modified = _SPIRE_WARN_RE.sub(b'', modified)
            for name in img_names:
                modified = re.sub(rb'/' + re.escape(name) + rb'\s+Do\b', b'', modified)
            if modified != raw:
                doc.update_stream(xref, modified)
        if os.path.exists(_LOGO_PATH):
            page.insert_image(fitz.Rect(*_LOGO_RECT_COORDS), filename=_LOGO_PATH, keep_proportion=True)
    doc.save(tmp_path, garbage=4, deflate=True)
    doc.close()
    del doc
    gc.collect()
    os.replace(tmp_path, pdf_path)


def list_projects(path_source, year, month, eligibility_rules, to_isolate_list, dict_partner_rename, tasks, df=None):
    # `df` lets the caller pass an already-parsed DataFrame so the CSV is read
    # and parsed once even when listing both Peve and Fausto. We copy it because
    # filter_employee_and_partners mutates the frame in place.
    df = load_df(path_source, year, month) if df is None else df.copy()
    if df is None or df.empty:
        return []
    df = filter_employee_and_partners(df, eligibility_rules, dict_partner_rename, to_isolate_list)
    df = identify_suspicious_data(df)
    df = df[df['flag_suspicious'] == False].copy()
    result = []
    for task_category in tasks:
        pp = df[df['task_category'] == task_category]\
            .groupby(['task_category', 'partner_name', 'project_name'])\
            .agg({'employee_name': lambda x: list(x)})\
            .reset_index()
        for _, row in pp.iterrows():
            result.append({
                'task_category': row['task_category'],
                'partner_name': row['partner_name'],
                'project_name': row['project_name'],
            })
    return result


def create_rapportini(path_source, path_output, year, month, filtered_partners, eligibility_rules, to_isolate_list, dict_partner_rename, tasks, only_task=None, only_partner=None, only_project=None):
    import spire.xls as spire  # heavy .NET CLR; imported here so listing never loads it
    year = int(year)
    month = int(month)

    print('Pre - create folders')
    create_folders(path_output, str(year), str(month),  'Fausto')
    print('Post - create folders')

    out_dir = path_output + str(year) + '_' + str(month) + '/'

    memlog.snapshot('fausto create_rapportini: before load_df')
    df_SOURCE = load_df(path_source, year, month)
    memlog.snapshot(f'fausto: after load_df (rows={0 if df_SOURCE is None else len(df_SOURCE)})')
    df_all = filter_employee_and_partners(df_SOURCE, eligibility_rules, dict_partner_rename, to_isolate_list)
    df_all = identify_suspicious_data(df_all)
    df_all = df_all[df_all['flag_suspicious'] == False].copy()
    memlog.snapshot(f'fausto: after filters (rows={len(df_all)})')
    print(f'df_all after suspicious filter: {df_all.shape}')
    
    for task_category in ['Assistenza', 'Intervento']:
        if only_task and task_category != only_task:
            continue
        print('')
        print('')
        print(task_category)

        partners_projects = df_all[df_all['task_category'] == task_category]\
            .groupby(['task_category','partner_name','project_name'])\
            .agg({'employee_name': lambda x: list(x)})\
            .reset_index()    
        print('df partner_projects: ',partners_projects)
        
        template_name = 'templates/Template - Rapportino - Fausto.xlsx'
        print(template_name)

        #partners_projects = partners_projects[(partners_projects['partner_name']=='Icmi Srl') | (partners_projects['partner_name']=='F.lli Vaser Snc')]

        rows_to_process = partners_projects.sort_values(['partner_name'])
        total_to_process = len(rows_to_process)
        progress_done = 0
        for index, row in rows_to_process.iterrows():
            partner = row['partner_name']
            project = row['project_name']
            if only_partner and partner != only_partner:
                continue
            if only_project is not None and project != only_project:
                continue
            employees = list(set(row['employee_name']))
            flag_to_isolate = partner in to_isolate_list
            progress_done += 1
            # Sentinel line the backend turns into a "Elaborazione N/M" counter.
            print(f'@@PROGRESS@@\x1f{progress_done}\x1f{total_to_process}\x1f{task_category} · {partner} - {project}')
            print(f'Inizio {task_category} - {partner} - {project} ')
            print('df pre filter: ',df_all.shape)
            memlog.snapshot(f'fausto [{progress_done}/{total_to_process}] START {task_category} {partner} / {project}')

            wb = load_workbook(template_name)
            ws = wb.active
            name_file_riassunto = str(year) + '-' + str(month) + '_' + task_category + '.xlsx'
            print('Post - load templates: ', name_file_riassunto)    

            df = filter_df(df = df_all,
                           employees = employees,
                           partner_name = partner,
                           project_name = project,
                           task_category = task_category,
                           flag_to_isolate = flag_to_isolate)
            
            df['name_clean'] = df.apply(lambda x: clean_description(x,'x_studio_titolo'), axis=1)
            print(df.shape)

            df_ = df.groupby(['employee_name', 'date', 'task_name', 'project_name_orig']).agg(
                 {'duration_unit_amount': 'sum',
                  'name_clean': lambda x: list(x),
                  'time_period': lambda x: list(x)
                  }).reset_index()
            print(df_.shape)

            ws = wb.active
            wb.copy_worksheet(ws)
            print('worksheet copied')
            ws = wb.worksheets[-1]
            print('template sheet removed')
            ws.sheet_view.showGridLines = False
            ws.title = (re.split(r"[ ]", partner)[0] + ('-' if project != '' else '') + project)[:31]
            print('title and sheet view changed')

            row_start = 13
            j = row_start
            
            for row in ['B3', 'B35', 'B67', 'B99', 'B131', 'B163', 'B195', 'B227', 'B259', 'B291', 'B323']:
                ws[row].value = partner 
                ws[row].font = openpyxl.styles.Font(name='Arial', size=14, bold=True)
            print('Partner set')

            for row in ['G3', 'G35', 'G67', 'G99', 'G131', 'G163', 'G195', 'G227', 'G259', 'G291', 'G323']:
                ws[row].value = str(datetime(year, month, cal.monthrange(year, month)[1]).strftime('%d/%m/%y'))
                ws[row].font = openpyxl.styles.Font(name='Arial', size=12, bold=True)
            print('Calendar set')   
 
            id_breaks = [32, 64, 96, 128, 160, 192, 224, 256, 288, 320]
            max_id = 0
            for id in range(0, len(id_breaks)):
                id_break = id_breaks[id]
                row_break = Break(id=id_break)
                ws.row_breaks.append(row_break)
                max_id = id
            print('Break set') 

            prev_employee = ''    
            num_sheet = 1
            df_pages = pd.DataFrame()
            row_starts = [0, 14, 46, 78, 110, 144, 174, 206, 238, 270, 302]
            row_ends = [0, 28, 60, 92, 124, 156, 188, 220, 252, 284]
            
            for index, row in df_.iterrows():
                print(row)

                data = row['date'].strftime('%d-%m')
                task_name = row['task_name']
                name_clean = row['name_clean']
                time_period = row['time_period']
                duration_unit_amount = row['duration_unit_amount']
                employee = row['employee_name']
                
                if task_category == 'Assistenza':
                    project_name_orig_str = (' / ' + row['project_name_orig'])
                else:
                    project_name_orig_str = '' 
                #print('    ---project_name_orig done')

                j, num_sheet, df_pages = next_row_num(
                    j=j, 
                    num_sheet=num_sheet, 
                    row_starts=row_starts, 
                    row_ends=row_ends, 
                    prev_employee=prev_employee, 
                    employee=employee, 
                    df_pages=df_pages)
                prev_employee = employee
                print('func next_row_num executed') 

                ws['B'+str(j)].value = data
                ws['B'+str(j)].font = openpyxl.styles.Font(bold=True)
                ws['B'+str(j)].border = openpyxl.styles.borders.Border(left = openpyxl.styles.borders.Side(border_style='double', color='FF000000'))
                if task_name == 'Viaggio':
                    ws['C'+str(j)].value = duration_unit_amount
                    ws['C'+str(j)].font = openpyxl.styles.Font(bold=True)
                else:
                    ws['D'+str(j)].value = duration_unit_amount 
                    ws['D'+str(j)].font = openpyxl.styles.Font(bold=True)
                ws['H'+str(j)].value = task_name + ': ' + str(time_period).replace('[', '').replace(']', '').replace("'", '') + project_name_orig_str
                ws['H'+str(j)].font = openpyxl.styles.Font(bold=True)
                 
                if len(name_clean) > 1:
                    name_clean_unique = [x for x in name_clean if x != '']
                    used = set()
                    name_clean_unique = [x for x in name_clean_unique if x not in used and (used.add(x) or True)]
                else:
                    name_clean_unique = name_clean

                for i in range(0, len(name_clean_unique)):
                    j, num_sheet, df_pages = next_row_num(
                        j=j, 
                        num_sheet=num_sheet, 
                        row_starts=row_starts, 
                        row_ends=row_ends, 
                        prev_employee=prev_employee, 
                        employee=employee, 
                        df_pages=df_pages)
                    ws['H'+str(j)].value = name_clean_unique[i].capitalize()
                    ws['B'+str(j)].border = openpyxl.styles.borders.Border(left = Side(border_style='double', color='FF000000'),
                                                right = Side(border_style='thin', color='FF000000'),
                                                bottom = Side(border_style='thin', color='FF000000'),
                                                top = Side(border_style='thin', color='FF000000'))
             
                print('data, start, stop, name_clean set in file') 
    
            row_employee_name = [30, 62, 94, 126, 158, 190, 222, 254, 286, 316]
            for index, row in df_pages.iterrows():
                row_num = 'H{}'.format(str(row_employee_name[row['page']-1]))
                ws[row_num].value = row['employee']        
            print('employee name set') 

            max_page = df_pages['page'].max()
            ws.print_area = 'A1:V{}'.format(row_ends[max_page]+4)
            ws.delete_rows(row_ends[max_page]+4, 1000)
            print('print area set')

            if os.path.exists(_LOGO_PATH):
                _m = wb.worksheets[0].page_margins
                _left_cm = (_m.left   or 0.70) * 2.54
                _top_cm  = (_m.top    or 0.75) * 2.54
                _bot_cm  = (_m.bottom or 0.75) * 2.54
                _ANC_X  = _left_cm + 18.2
                _ANC_Y  = _top_cm  + (-1.3)
                _L_W    = 53  / 72 * 2.54
                _L_H    = 40  / 72 * 2.54
                _PAGE_H = 17.92  # calibrated page step (row heights of one page)
                for page_idx in range(1, int(max_page) + 1):
                    logo_img = Image(_LOGO_PATH)
                    logo_img.anchor = AbsoluteAnchor(
                        pos=XDRPoint2D(
                            cm_to_EMU(_ANC_X),
                            cm_to_EMU(_ANC_Y + (page_idx - 1) * _PAGE_H),
                        ),
                        ext=XDRPositiveSize2D(cm_to_EMU(_L_W), cm_to_EMU(_L_H)),
                    )
                    ws.add_image(logo_img)

            print(f'Fine {task_category} - {partner} - {project} - {employee} - {j} - {row_ends[max_page]}')

            wb.remove(wb.worksheets[0])
            print('sheet removed')
            xlsx_name = (str(year) + '-' + str(month) + ' _' + task_category.upper() + '_'
                         + re.split(r"[ ]", partner)[0] + (' - ' if project != '' else '') + project + '.xlsx')
            xlsx_path = os.path.abspath(os.path.join(out_dir, xlsx_name))
            wb.save(xlsx_path)
            print('sheet saved')
            wb.close()
            del wb
            print('sheet closed')
            memlog.snapshot(f'fausto [{progress_done}/{total_to_process}] xlsx saved+closed, before Spire PDF')

            # Convert this partner's workbook to PDF right away, then release
            # everything before moving to the next partner. Spire loads the whole
            # workbook into the .NET heap, so we Dispose() + gc.collect() per file
            # to keep peak memory flat (Render free tier is 512MB). Generating and
            # converting one partner at a time avoids holding every file in memory.
            pdf_path = xlsx_path.replace('.xlsx', '.pdf')
            print('Generating PDF locally...')
            try:
                wb_pdf = spire.Workbook()
                wb_pdf.LoadFromFile(xlsx_path)
                memlog.snapshot(f'fausto [{progress_done}/{total_to_process}] Spire LoadFromFile done')
                wb_pdf.SaveToFile(pdf_path, spire.FileFormat.PDF)
                memlog.snapshot(f'fausto [{progress_done}/{total_to_process}] Spire SaveToFile(PDF) done')
                wb_pdf.Dispose()
                del wb_pdf
                _clean_spire_pdf(pdf_path)
                print(f'PDF generated: {pdf_path}')
                try:
                    from storage import upload_and_remove, to_storage_key
                    # Upload then delete the local copies: /tmp is RAM-backed
                    # tmpfs on Render and counts against the 512MB cap.
                    upload_and_remove(xlsx_path, to_storage_key(xlsx_path))
                    upload_and_remove(pdf_path,  to_storage_key(pdf_path))
                    print(f'Uploaded to Supabase: {to_storage_key(xlsx_path)}')
                except Exception as sup_err:
                    print(f'Supabase upload failed (non-fatal): {sup_err}')
            except Exception as spire_err:
                print(f'Spire PDF failed for {xlsx_name}, skipping: {spire_err}')
            finally:
                gc.collect()
                memlog.snapshot(f'fausto [{progress_done}/{total_to_process}] END (after Dispose+gc.collect)')

    print(glob.glob('*'))
    print(glob.glob(path_output+'/*'))
    print([i for i in os.listdir(out_dir) if i.endswith(".xlsx")])
    print('PDF generation complete')
