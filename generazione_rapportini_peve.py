import numpy as np
from math import ceil
import pandas as pd
import re
from datetime import datetime
from datetime import timedelta
import calendar as cal

import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.cell import MergedCell
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Color, PatternFill, Alignment
from openpyxl.worksheet.pagebreak import Break
from openpyxl.drawing.image import Image
from openpyxl.styles.borders import Border, Side
from openpyxl.drawing.spreadsheet_drawing import AbsoluteAnchor
from openpyxl.drawing.xdr import XDRPoint2D, XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU, cm_to_EMU

import spire.xls as spire
from spire.xls import *
from spire.xls import ExcelVersion
import fitz  # PyMuPDF
import re

import os
from os import listdir
from os.path import isfile, join
import string
import shutil
import glob
import gc

import memlog


def is_daylight(date):
    start = datetime.strptime("31-03-2024", "%d-%m-%Y")
    end = datetime.strptime("27-10-2024", "%d-%m-%Y")
    if start <= date <= end:
        return 2
    else:
        return 1


def load_df(path_output, year, month):
    print(f"--- Starting load_df for {year}-{month} ---")
    file_path = path_output + str(year) + '_' + str(month) + '_timesheets_extraction.csv'

    try:
        df = pd.read_csv(file_path)
        df.columns = df.columns.str.strip()
        print(f"DEBUG: Loaded {len(df)} rows.")
    except Exception as e:
        print(f"!!! Error reading CSV: {e}")
        return None

    df['date'] = pd.to_datetime(df['date'], format='%Y-%m-%d')
    df['year'] = df['date'].dt.year
    df['month'] = df['date'].dt.month

    df['time_diff'] = df['date'].apply(lambda x: is_daylight(x))

    print("DEBUG: Converting date_start/stop strings to datetime...")
    df['date_start'] = pd.to_datetime(df['date_start']) + pd.to_timedelta(df['time_diff'], unit='h')
    df['date_stop'] = pd.to_datetime(df['date_stop']) + pd.to_timedelta(df['time_diff'], unit='h')

    df['hour_start'] = df['date_start'].dt.strftime('%H:%M')
    df['hour_stop'] = df['date_stop'].dt.strftime('%H:%M')
    df['time_period'] = df['hour_start'] + '-' + df['hour_stop']

    def get_match(val, pattern):
        match = re.search(pattern, str(val))
        return match.group(1) if match else "Unknown"

    print("DEBUG: Extracting names via Regex...")
    df['employee_name'] = df['employee_id'].apply(lambda x: get_match(x, ", '(.*) - "))
    df['partner_name'] = df['partner_id'].apply(lambda x: string.capwords(get_match(x.replace('"', "'"), ", '(.*)']")))
    df['project_name_orig'] = df['project_id'].apply(lambda x: get_match(x, r" - (.*) \[",))

    print("DEBUG: Filtering tasks...")
    valid_tasks = ['Intervento', 'Viaggio', 'Assistenza']
    df = df[df['task_name'].isin(valid_tasks)].copy()

    if df.empty:
        print(f"!!! WARNING: No rows match tasks {valid_tasks}. Check for case-sensitivity.")
        return df

    df['task_category'] = df['task_name'].apply(lambda x: 'Assistenza' if x == 'Assistenza' else 'Intervento')

    df = df[(df['year'] == int(year)) & (df['month'] == int(month))]

    print(f"DEBUG: Rows remaining after year/month filter: {len(df)}")

    if not df.empty:
        df = df.sort_values('hour_start')

    print("--- load_df finished ---")
    return df


def hours_by_task_name(row, task_name):
    if row['task_name'] in task_name:
        return row['duration_unit_amount']
    else:
        return None


def clear_cells(ws):
    for row in ws['B7:I36']:
        for cell in row:
            cell.value = None
    for row in ws['N7:N36']:
        for cell in row:
            cell.value = None

    merged_cells_ranges = ws.merged_cells.ranges
    for merged in merged_cells_ranges:
        ws.unmerge_cells(str(merged))
    for row in ws['J7:J36']:
        for cell in row:
            cell.value = None
    for i in range(6, 37):
        ws.merge_cells('J' + str(i) + ':L' + str(i))

    return ws


def weekday_or_weekend(row):
    weekno = row['date'].weekday()
    if weekno < 5:
        return ''
    else:
        return 'F'


def clean_description(row, field):
    print(f'   --- func clean_description - Iniziato')
    print(f'   --- func clean_description - field:{field}')

    name = str(row[field])
    print(f'   --- func clean_description - name:{name}')

    name = name.replace('<p>.</p>', '')
    name = name.replace('</p>', '')
    name = name.replace('<p>', '')
    name = name.replace('&nbsp', '')
    name = name.replace('<br>', '')
    name = name.replace('<span style="width: 40px;" class="oe-tabs">\t</span>​', '')
    name = name.replace('<span style="width: 40px;" class="oe-tabs"> </span>​', '')
    name = name.replace('<span class="oe-tabs" style="width: 40px;">	</span>', '')
    print(f'   --- func clean_description - Finito')
    return name


def filter_df(df, employees, partner_name, project_name, task_category, flag_to_isolate):
    df = df[df['employee_name'].isin(employees)]
    if flag_to_isolate:
        df = df[(df['partner_name'] == partner_name) & (df['project_name'] == project_name) & (df['task_category'] == task_category)]
    else:
        df = df[(df['partner_name'] == partner_name) & (df['task_category'] == task_category)]
    df = df.sort_values(['task_name'], ascending=False)
    return df


def next_row_num(j, num_sheet, row_starts, row_ends, prev_employee, employee, df_pages):
    if j in row_ends:
        num_sheet += 1
        j = row_starts[num_sheet]
        df_page_new = pd.DataFrame({'page': [num_sheet], 'employee': [employee]})
        df_pages = pd.concat([df_pages, df_page_new], ignore_index=True)

    elif prev_employee == '':
        j += 1
        df_page_new = pd.DataFrame({'page': [num_sheet], 'employee': [employee]})
        df_pages = pd.concat([df_pages, df_page_new], ignore_index=True)

    elif prev_employee != employee:
        num_sheet += 1
        j = row_starts[num_sheet]
        df_page_new = pd.DataFrame({'page': [num_sheet], 'employee': [employee]})
        df_pages = pd.concat([df_pages, df_page_new], ignore_index=True)

    else:
        j += 1

    return j, num_sheet, df_pages


def create_folders(path, year, month, owner):
    path_output = path + year + '_' + month + '/'
    os.makedirs(path_output, exist_ok=True)


def filter_employee_and_partners(df, eligibility_rules, dict_partner_rename, to_isolate_list):
    df['project_name'] = df.apply(lambda x: x['project_name_orig'] if x['partner_name'] in to_isolate_list else '', axis=1)

    eligible_employees = [k for (k, v) in eligibility_rules.items()]
    df = df[df['employee_name'].isin(eligible_employees)]

    df['eligible_partners'] = df['employee_name'].map(eligibility_rules)
    df['eligibility'] = df.apply(lambda x: x['eligible_partners'] == ['*'] or (x['partner_name'] in x['eligible_partners']), axis=1)
    df = df[df['eligibility'] == True]

    df['partner_name'] = df['partner_name'].map(dict_partner_rename).fillna(df['partner_name'])

    return df


def identify_suspicious_data(df):
    flag_suspicious = []
    daterange = []

    for index, row in df.iterrows():
        project = row['project_name']
        days_to_check = [row['date'] - timedelta(days=1), row['date'], row['date'] + timedelta(days=1)]
        flag_suspicious += [('Intervento' not in df[(df['date'].isin(days_to_check)) & (df['partner_name'] == row['partner_name']) & (df['employee_name'] == row['employee_name'])]['task_name'].unique()) & (row['task_name'] == 'Viaggio')]
    df['flag_suspicious'] = flag_suspicious

    return df


_SPIRE_WARN_RE = re.compile(
    rb'BT /FAAA\w+ \S+ Tf \S+ \S+ \S+ \S+ \S+ \S+ Tm 1 0 0 rg \[.+?\] TJ\s+ET',
    re.DOTALL,
)

_LOGO_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'templates', 'logo_solware.png')
_LOGO_RECT = fitz.Rect(620, 77, 673, 117)

def _clean_spire_pdf(pdf_path):
    tmp_path = pdf_path + '.tmp'
    doc = fitz.open(pdf_path)
    if doc.page_count > 1:
        doc.delete_page(-1)
    for page in doc:
        img_names = {img[7].encode() for img in page.get_images(full=True)}
        for xref in page.get_contents():
            raw = doc.xref_stream(xref)
            modified = raw
            if b'1 0 0 rg' in modified:
                modified = _SPIRE_WARN_RE.sub(b'', modified)
            for name in img_names:
                modified = re.sub(rb'/' + re.escape(name) + rb'\s+Do\b', b'', modified)
            if modified != raw:
                doc.update_stream(xref, modified)
        if os.path.exists(_LOGO_PATH):
            page.insert_image(_LOGO_RECT, filename=_LOGO_PATH, keep_proportion=True)
    doc.save(tmp_path, garbage=4, deflate=True)
    doc.close()
    del doc
    gc.collect()
    os.replace(tmp_path, pdf_path)


def list_projects(path_source, year, month, eligibility_rules, to_isolate_list, dict_partner_rename, tasks, df=None):
    # `df` lets the caller pass an already-parsed DataFrame so the CSV is read
    # and parsed once even when listing both Peve and Fausto. We copy it because
    # filter_employee_and_partners mutates the frame in place.
    df = load_df(path_source, year, month) if df is None else df.copy()
    if df is None or df.empty:
        return []
    df = filter_employee_and_partners(df, eligibility_rules, dict_partner_rename, to_isolate_list)
    df = identify_suspicious_data(df)
    df = df[df['flag_suspicious'] == False].copy()
    result = []
    for task_category in tasks:
        pp = df[df['task_category'] == task_category]\
            .groupby(['task_category', 'partner_name', 'project_name'])\
            .agg({'employee_name': lambda x: list(x)})\
            .reset_index()
        for _, row in pp.iterrows():
            result.append({
                'task_category': row['task_category'],
                'partner_name': row['partner_name'],
                'project_name': row['project_name'],
            })
    return result


def create_rapportini(path_source, path_output, year, month, filtered_partners, eligibility_rules, to_isolate_list, dict_partner_rename, tasks, only_task=None, only_partner=None, only_project=None):
    year = int(year)
    month = int(month)

    print('Pre - create folders')
    create_folders(path_output, str(year), str(month), 'Peve')
    print('Post - create folders')

    out_dir = path_output + str(year) + '_' + str(month) + '/'

    memlog.snapshot('peve create_rapportini: before load_df')
    df_SOURCE = load_df(path_source, year, month)
    memlog.snapshot(f'peve: after load_df (rows={0 if df_SOURCE is None else len(df_SOURCE)})')
    df_all = filter_employee_and_partners(df_SOURCE, eligibility_rules, dict_partner_rename, to_isolate_list)
    memlog.snapshot(f'peve: after filter_employee_and_partners (rows={len(df_all)})')
    print(df_all.shape)

    for task_category in ['Assistenza', ]:
        if only_task and task_category != only_task:
            continue
        print('')
        print('')
        print(task_category)

        partners_projects = df_all[df_all['task_category'] == task_category]\
            .groupby(['task_category', 'partner_name', 'project_name'])\
            .agg({'employee_name': lambda x: list(x)})\
            .reset_index()
        print('df partner_projects: ', partners_projects)

        template_name = 'templates/Template - Rapportino - Peve.xlsx'
        print(template_name)

        rows_to_process = partners_projects.sort_values(['partner_name'])
        total_to_process = len(rows_to_process)
        progress_done = 0
        for index, row in rows_to_process.iterrows():
            partner = row['partner_name']
            project = row['project_name']
            if only_partner and partner != only_partner:
                continue
            if only_project is not None and project != only_project:
                continue
            employees = list(set(row['employee_name']))
            flag_to_isolate = partner in to_isolate_list
            progress_done += 1
            # Sentinel line the backend turns into a "Elaborazione N/M" counter.
            print(f'@@PROGRESS@@\x1f{progress_done}\x1f{total_to_process}\x1f{task_category} · {partner} - {project}')
            print(f'{task_category} - {partner} - {project} - {flag_to_isolate} - Inizio elaborazione')
            memlog.snapshot(f'peve [{progress_done}/{total_to_process}] START {partner} / {project}')

            wb = load_workbook(template_name)
            ws = wb.active
            name_file_riassunto = str(year) + '-' + str(month) + '_' + task_category + '.xlsx'
            print(f'{task_category} - {partner} - {project} - {flag_to_isolate} - Template caricato e name_file_riassunto impostato')

            df = filter_df(df=df_all,
                           employees=employees,
                           partner_name=partner,
                           project_name=project,
                           task_category=task_category,
                           flag_to_isolate=flag_to_isolate)
            print(f'{task_category} - {partner} - {project} - {flag_to_isolate} - Dataframe filtrato')

            df['name_clean'] = df.apply(lambda x: clean_description(x, 'x_studio_titolo'), axis=1)
            print(f'{task_category} - {partner} - {project} - {flag_to_isolate} - Func clean_description effettuata')

            df_ = df.loc[:, [
                'employee_name',
                'date',
                'task_name',
                'project_name_orig',
                'duration_unit_amount',
                'name_clean',
                'hour_start',
                'hour_stop'
            ]].sort_values(['employee_name', 'date'])
            print(f'{task_category} - {partner} - {project} - {flag_to_isolate} - Func sort_values effettuata')

            ws = wb.active
            wb.copy_worksheet(ws)
            ws = wb.worksheets[-1]
            ws.sheet_view.showGridLines = False
            ws.title = (re.split(r"[ ]", partner)[0] + ('-' if project != '' else '') + project)[:31]
            print(f'{task_category} - {partner} - {project} - {flag_to_isolate} - Sheet copiato e titolo cambiato')

            row_start = 13
            j = row_start

            for row in ['B3', 'B35', 'B67', 'B99', 'B131', 'B163', 'B195', 'B227', 'B259', 'B291', 'B323']:
                ws[row].value = partner
                ws[row].font = openpyxl.styles.Font(name='Arial', size=14, bold=True)
            print(f'{task_category} - {partner} - {project} - {flag_to_isolate} - Nome partner impostato')

            for row in ['G3', 'G35', 'G67', 'G99', 'G131', 'G163', 'G195', 'G227', 'G259', 'G291', 'G323']:
                ws[row].value = str(datetime(year, month, cal.monthrange(year, month)[1]).strftime('%d/%m/%y'))
                ws[row].font = openpyxl.styles.Font(name='Arial', size=12, bold=True)
            print(f'{task_category} - {partner} - {project} - {flag_to_isolate} - Data impostata')

            for row in ['J7', 'J39', 'J71', 'J103', 'J135', 'J167', 'J199', 'J231', 'J263', 'J295', 'J327']:
                ws[row].value = project
            print(f'{task_category} - {partner} - {project} - {flag_to_isolate} - Nome progetto impostato')

            id_breaks = [32, 64, 96, 128, 160, 192, 224, 256, 288, 320]
            max_id = 0
            for id in range(0, len(id_breaks)):
                id_break = id_breaks[id]
                row_break = Break(id=id_break)
                ws.row_breaks.append(row_break)
                max_id = id
            print(f'{task_category} - {partner} - {project} - {flag_to_isolate} - Break lines impostate')

            prev_employee = ''
            num_sheet = 1
            df_pages = pd.DataFrame()
            row_starts = [0, 14, 46, 78, 110, 144, 174, 206, 238, 270, 302]
            row_ends = [0, 28, 60, 92, 124, 156, 188, 220, 252, 284]

            for index, row in df_.iterrows():
                data = row['date'].strftime('%d-%m')
                task_name = row['task_name']
                name_clean = row['name_clean'].capitalize()
                hour_start = row['hour_start']
                hour_stop = row['hour_stop']
                duration_unit_amount = row['duration_unit_amount']
                employee = row['employee_name']

                if task_category == 'Assistenza':
                    project_name_orig_str = (' / ' + row['project_name_orig'])
                else:
                    project_name_orig_str = ''

                j, num_sheet, df_pages = next_row_num(
                    j=j,
                    num_sheet=num_sheet,
                    row_starts=row_starts,
                    row_ends=row_ends,
                    prev_employee=prev_employee,
                    employee=employee,
                    df_pages=df_pages)
                prev_employee = employee

                ws['B' + str(j)].value = data
                ws['D' + str(j)].value = hour_start
                ws['E' + str(j)].value = hour_stop
                ws['H' + str(j)].value = name_clean
            print(f'{task_category} - {partner} - {project} - {flag_to_isolate} - Righe della tabella popolate')

            row_employee_name = [30, 62, 94, 126, 158, 190, 222, 254, 286, 316]
            for index, row in df_pages.iterrows():
                row_num = 'H{}'.format(str(row_employee_name[int(row['page']) - 1]))
                ws[row_num].value = row['employee']
            print(f'{task_category} - {partner} - {project} - {flag_to_isolate} - Nome del dipendente popolato')

            max_page = df_pages['page'].max()
            ws.print_area = 'A1:V{}'.format(row_ends[int(max_page)] + 4)
            ws.delete_rows(row_ends[int(max_page)] + 4, 1000)

            if os.path.exists(_LOGO_PATH):
                _m = wb.worksheets[0].page_margins
                _left_cm = (_m.left   or 0.70) * 2.54
                _top_cm  = (_m.top    or 0.75) * 2.54
                _bot_cm  = (_m.bottom or 0.75) * 2.54
                _ANC_X  = _left_cm + 18.2
                _ANC_Y  = _top_cm  + (-1.3)
                _L_W    = 53  / 72 * 2.54
                _L_H    = 40  / 72 * 2.54
                _PAGE_H = 17.92  # calibrated page step (row heights of one page)
                for page_idx in range(1, int(max_page) + 1):
                    logo_img = Image(_LOGO_PATH)
                    logo_img.anchor = AbsoluteAnchor(
                        pos=XDRPoint2D(
                            cm_to_EMU(_ANC_X),
                            cm_to_EMU(_ANC_Y + (page_idx - 1) * _PAGE_H),
                        ),
                        ext=XDRPositiveSize2D(cm_to_EMU(_L_W), cm_to_EMU(_L_H)),
                    )
                    ws.add_image(logo_img)

            print(f'{task_category} - {partner} - {project} - {employee} - {j} - {row_ends[int(max_page)]} - Fine elaborazione')

            wb.remove(wb.worksheets[0])
            print('sheet removed')
            xlsx_name = (str(year) + '-' + str(month) + ' _' + task_category.upper() + '_'
                         + re.split(r"[ ]", partner)[0] + (' - ' if project != '' else '') + project + '.xlsx')
            xlsx_path = os.path.abspath(os.path.join(out_dir, xlsx_name))
            wb.save(xlsx_path)
            print('sheet saved')
            wb.close()
            del wb
            print('sheet closed')
            memlog.snapshot(f'peve [{progress_done}/{total_to_process}] xlsx saved+closed, before Spire PDF')

            # Convert this partner's workbook to PDF right away, then release
            # everything before moving to the next partner. Spire loads the whole
            # workbook into the .NET heap, so we Dispose() + gc.collect() per file
            # to keep peak memory flat (Render free tier is 512MB). Generating and
            # converting one partner at a time avoids holding every file in memory.
            pdf_path = xlsx_path.replace('.xlsx', '.pdf')
            print('Generating PDF locally...')
            try:
                wb_pdf = spire.Workbook()
                wb_pdf.LoadFromFile(xlsx_path)
                memlog.snapshot(f'peve [{progress_done}/{total_to_process}] Spire LoadFromFile done')
                wb_pdf.SaveToFile(pdf_path, spire.FileFormat.PDF)
                memlog.snapshot(f'peve [{progress_done}/{total_to_process}] Spire SaveToFile(PDF) done')
                wb_pdf.Dispose()
                del wb_pdf
                _clean_spire_pdf(pdf_path)
                print(f'PDF generated: {pdf_path}')
                try:
                    from storage import upload_and_remove, to_storage_key
                    # Upload then delete the local copies: /tmp is RAM-backed
                    # tmpfs on Render and counts against the 512MB cap.
                    upload_and_remove(xlsx_path, to_storage_key(xlsx_path))
                    upload_and_remove(pdf_path,  to_storage_key(pdf_path))
                    print(f'Uploaded to Supabase: {to_storage_key(xlsx_path)}')
                except Exception as sup_err:
                    print(f'Supabase upload failed (non-fatal): {sup_err}')
            except Exception as spire_err:
                print(f'Spire PDF failed for {xlsx_name}, skipping: {spire_err}')
            finally:
                gc.collect()
                memlog.snapshot(f'peve [{progress_done}/{total_to_process}] END (after Dispose+gc.collect)')

    print(glob.glob('*'))
    print(glob.glob(path_output + '/*'))
    print([i for i in os.listdir(out_dir) if i.endswith(".xlsx")])
    print('PDF generation complete')
