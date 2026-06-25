import numpy as np
from math import ceil
import pandas as pd
import re
from datetime import datetime
from datetime import timedelta
import calendar as cal

from openpyxl import Workbook, load_workbook
from openpyxl.cell import MergedCell
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Color, PatternFill, Alignment
from openpyxl.worksheet.pagebreak import Break
from openpyxl.drawing.image import Image
from openpyxl.styles.borders import Border, Side
from openpyxl.drawing.spreadsheet_drawing import AbsoluteAnchor
from openpyxl.drawing.xdr import XDRPoint2D, XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU, cm_to_EMU

import spire.xls as spire
from spire.xls import ExcelVersion

import os
from os import listdir
from os.path import isfile, join
import string
import shutil
import glob
import gc


def is_daylight(date):
    start = datetime.strptime("31-03-2024", "%d-%m-%Y")
    end = datetime.strptime("27-10-2024", "%d-%m-%Y")
    if start <= date <= end:
        return 2
    else:
        return 1


def load_df(path_output, eligibility_rules, year, month):
    print('starting the load_df function')
    print(glob.glob('*'))
    print(glob.glob('Odoo exports/*'))
    print(glob.glob('Odoo exports/2024_8/*'))
    print(path_output)
    file_path = path_output+str(year)+'_'+str(month)+'_timesheets_extraction.csv' 
    print('file to load: ', file_path)

    df = pd.read_csv(file_path)
    print('df loaded by timesheets_extraction')
    print(df)

    try:
        df['date'] = pd.to_datetime(df['date'], format='%Y-%m-%d')
        print('... Estrazione data fatta')
        df['year'] = df['date'].apply(lambda x: x.strftime("%Y")).astype(int)
        print('... Estrazione anno fatta')
        df['month'] = df['date'].apply(lambda x: x.strftime("%m")).astype(int)
        print('... Estrazione mese fatta')
        df['day'] = df['date'].apply(lambda x: x.strftime("%d")).astype(int)
        print('... Estrazione giorno fatta')

        df['time_diff'] = df['date'].apply(lambda x: is_daylight(x))
        df['date_start'] =  df.loc[:,['date_start','time_diff']].apply(
            lambda x: datetime.strptime(x['date_start'], '%Y-%m-%d %H:%M:%S') + timedelta(hours=x['time_diff'])
        , axis=1)
        df['date_stop'] = df.loc[:,['date_stop','time_diff']].apply(
            lambda x: datetime.strptime(x['date_stop'], '%Y-%m-%d %H:%M:%S') + timedelta(hours=x['time_diff'])
        , axis=1)
        print('... Calcolo del time_diff, data_start e date_stop fatta')

        df['employee_name'] = df['employee_id'].apply(lambda x: re.search(", '(.*) - ", x).group(1))
        df = filter_employees(df, eligibility_rules)
        
        df['partner_name'] = df['partner_id'].apply(lambda x: string.capwords(re.search(", '(.*)']", x.replace('"',"'")).group(1)))
        df['project_name_orig'] = df['project_id'].apply(lambda x: re.search(r" - (.*) \[", x).group(1))
        print('... Estrazione employee, partner e project name fatta')
    
        df['hour_start'] = df.apply(lambda x: x['date_start'].strftime('%H:%M'), axis=1)
        df['hour_stop'] = df.apply(lambda x: x['date_stop'].strftime('%H:%M'), axis=1)
        df['time_period'] = df['hour_start'] + '-' + df['hour_stop']
        print('... Calcolo di hour_start, hour_stop e time_period fatta')

        df = df[df['task_name'].isin(['Intervento', 'Viaggio', 'Assistenza'])]
        df['task_category'] = df['task_name'].apply(lambda x: 'Assistenza' if x=='Assistenza' else 'Intervento')
        print('... Calcolo del task_category fatta')
        
        df['ore_lavoro'] = df.apply(lambda x: hours_by_task_name(x, ['Assistenza','Intervento']), axis=1)
        df['ore_viaggio'] = df.apply(lambda x: hours_by_task_name(x, 'Viaggio'), axis=1)
        df['tipo_giorno'] = df.apply(lambda x: weekday_or_weekend(x), axis=1)    
        print('df all transformed')
        print(df)

        df = df[(df['year'] == int(year)) & (df['month'] == int(month))]
        df = df.sort_values('hour_start')
        print('df filtered')
        print(df)

    except Exception as e:
        print(f"Error in load_df: ", e)
        raise

    return df


def filter_employees(df, eligibility_rules):

    eligible_employees = [k for (k, v) in eligibility_rules.items()]
    df = df[df['employee_name'].isin(eligible_employees)]
    print('df post employee filter: ',df.shape)
    
    return df


def filter_partners(df, eligibility_rules, dict_partner_rename, to_isolate_list):
    df['project_name'] = df.apply(lambda x: x['project_name_orig'] if x['partner_name'] in to_isolate_list else '', axis=1)
    
    df['eligible_partners'] = df['employee_name'].map(eligibility_rules)
    df['eligibility'] = df.apply(lambda x: x['eligible_partners']==['*'] or (x['partner_name'] in x['eligible_partners']), axis=1)
    df = df[df['eligibility']==True]
    print('df post eligibily filter: ',df.shape)
    
    df['partner_name'] = df['partner_name'].map(dict_partner_rename).fillna(df['partner_name'])
    print('df post partner mapping: ',df.shape)

    return df
    

def hours_by_task_name(row, task_name):
    if row['task_name'] in task_name:
        return row['duration_unit_amount']
    else:
        return None


def clear_cells(ws):
    for row in ws['B7:I36']:
        for cell in row:
            cell.value = None
    for row in ws['N7:N36']:
        for cell in row:
            cell.value = None

    merged_cells_ranges = ws.merged_cells.ranges
    for merged in merged_cells_ranges:
        ws.unmerge_cells(str(merged))
    for row in ws['J7:J36']:
        for cell in row:
            cell.value = None
    for i in range(6, 37):
        ws.merge_cells('J' + str(i) + ':L' + str(i))

    return ws


def weekday_or_weekend(row):
    weekno = row['date'].weekday()
    if weekno < 5:
        return ''
    else:
        return 'F'


def clean_description(row, field):
    name = row[field]
    name = name.replace('<p>.</p>', '')
    name = name.replace('</p>', '')
    name = name.replace('<p>', '')
    name = name.replace('&nbsp', '')
    name = name.replace('<br>', '')
    name = name.replace('<span style="width: 40px;" class="oe-tabs">\t</span>\u200b', '')
    name = name.replace('<span style="width: 40px;" class="oe-tabs"> </span>\u200b', '')
    name = name.replace('<span class="oe-tabs" style="width: 40px;">	</span>', '')
    return name


def filter_df(df, employees, partner_name, project_name, task_category, flag_to_isolate):
    df = df[df['employee_name'].isin(employees)]
    if flag_to_isolate:
        df = df[(df['partner_name'] == partner_name) & (df['project_name'] == project_name) & (df['task_category'] == task_category)]
    else:
        df = df[(df['partner_name'] == partner_name) & (df['task_category'] == task_category)]
    df = df.sort_values(['task_name'], ascending = False)
    return df


def next_row_num(j, num_sheet, row_starts, row_ends, prev_employee, employee, df_pages):
    if j in row_ends:
        num_sheet += 1
        j = row_starts[num_sheet]
        df_pages = df_pages.append({
                'page': num_sheet,
                'employee': employee
            }, ignore_index = True)
    elif prev_employee == '':
        j += 1
        df_pages = df_pages.append({
                'page': num_sheet,
                'employee': employee
            }, ignore_index = True)
    elif prev_employee != employee:
        num_sheet += 1
        j = row_starts[num_sheet]
        df_pages = df_pages.append({
                'page': num_sheet,
                'employee': employee
            }, ignore_index = True)
    else:
        j += 1
    
    return j, num_sheet, df_pages


def create_folders(path, year, month, owner):
    os.makedirs(path, exist_ok=True)
    path_output = path + year + '_' + month + '/'
    os.makedirs(path_output, exist_ok=True)


def identify_suspicious_data(df):
    flag_suspicious = []
    daterange = []
    
    for index, row in df.iterrows():
        project = row['project_name']
        days_to_check = [row['date'] - timedelta(days=1), row['date'], row['date'] + timedelta(days=1)]
        flag_suspicious += [('Intervento' not in df[(df['date'].isin(days_to_check)) & (df['partner_name'] == row['partner_name']) & (df['employee_name'] == row['employee_name'])]['task_name'].unique()) & (row['task_name'] == 'Viaggio')] 
    df['flag_suspicious'] = flag_suspicious
    
    return df


def create_riassunto(path_source, path_output, year, month, filtered_partners, eligibility_rules, to_isolate_list, dict_partner_rename, tasks):
    
    try:
        year = int(year)
        month = int(month)

        print('Pre - create folders')
        #path_output = path_output + str(year) + '_' + str(month) + '/'
        create_folders(path_output, str(year), str(month),  'Fausto')
        print('Post - create folders')
        
        df_SOURCE = load_df(path_source, eligibility_rules, year, month)
        df_all = filter_partners(df_SOURCE, eligibility_rules, dict_partner_rename, to_isolate_list)
        #df_all = filter_employee_and_partners(df_SOURCE, eligibility_rules, dict_partner_rename, to_isolate_list)  
        print(df_all.shape)
        
        for task_category in ['Assistenza', 'Intervento']:
            print('')
            print('')
            print(task_category)
        
            #print('df pre partner filter: ',df.shape)
            #if len(filtered_partners) > 0:
            #    df_all = df_all[df_all['partner_name'].isin(filtered_partners)]
            #print('df post partner filter: ',df.shape)
            
            partners_projects = df_all[df_all['task_category'] == task_category]\
                .groupby(['task_category','partner_name','project_name'])\
                .agg({'employee_name': lambda x: list(x)})\
                .reset_index()    
            print('df partner_projects: ',partners_projects)
            
            template_name = 'templates/Template - Riassunto.xlsx'
            print(template_name)
            
            wb = load_workbook(template_name)
            ws = wb.active
            name_file_riassunto = str(year) + '-' + str(month) + '_' + task_category + '.xlsx'
            print('Post - load templates: ', name_file_riassunto)          


            rows_to_process = partners_projects.sort_values(['partner_name'])
            total_to_process = len(rows_to_process)
            progress_done = 0
            for index, row in rows_to_process.iterrows():
                partner = row['partner_name']
                project = row['project_name']
                employees = list(set(row['employee_name']))
                flag_to_isolate = partner in to_isolate_list
                progress_done += 1
                # Sentinel line the backend turns into a "Elaborazione N/M" counter.
                print(f'@@PROGRESS@@\x1f{progress_done}\x1f{total_to_process}\x1f{task_category} · {partner} - {project}')
                print(f'Inizio {task_category} - {partner} - {project} ')
                print('df pre filter: ',df_all.shape)

                df = filter_df(df = df_all,
                            employees = employees,
                            partner_name = partner,
                            project_name = project,
                            task_category = task_category,
                            flag_to_isolate = flag_to_isolate)
                df = identify_suspicious_data(df)
                print('df post filter: ', df.shape)

                df_p = df.groupby(['date', 'employee_name', 'tipo_giorno']) \
                    .agg({
                        'duration_unit_amount': 'sum',
                        'ore_lavoro': 'sum',
                        'ore_viaggio': 'sum',
                        'flag_suspicious': 'max',
                        'task_name': lambda x:'/'.join(set(x))}) \
                    .reset_index()
                print('df_p: ', df_p)

                ws = wb.active
                wb.copy_worksheet(ws)
                print('worksheet copied')
                ws = wb.worksheets[-1]
                print('template sheet removed')
                ws.title = (re.split(r"[ ]", partner)[0] + ('-' if project != '' else '') + project)[:31]
                ws.sheet_view.showGridLines = False
                print('title and sheet view changed')

                cell_periodo_riferimento = "B2"
                print('cell_periodo_riferimento')
                cell_azienda = "B3"
                print('cell_azienda')
                cell_località = "B4"
                print('cell_località')

                print(year, ' - ', month, ' - ', task_category)
                print(year, ' - ', month, ' - ', cal.monthrange(int(year), int(month)), ' - ', task_category)
                ws[cell_periodo_riferimento].value = task_category+', ' + str(datetime(year, month, cal.monthrange(year, month)[1]).strftime('%d/%m/%y'))
                print('Periodo riferimento added to the sheet: ', cell_periodo_riferimento)
                ws[cell_azienda].value = 'Cliente: ' + partner
                print('Cliente added to the sheet')
                print('Flag_to_isolate: ', flag_to_isolate)
                ws[cell_località].value = ('Progetto: ' + project if flag_to_isolate is True else '')  
                print('Cell location added to the sheet')

                task_name_concat = ""
                for name in df_p['task_name'].unique():
                    task_name_concat = task_name_concat +' / '+ name
                print('Task name defined: ', task_name_concat)

                row_start = 7
                i = row_start
                for index, row in df_p.iterrows():
                    print('Row: ', row)
                    name = re.search("(.*) (.*)", row['employee_name']).group(1)[0]+'.'
                    surname = re.search("(.*) (.*)", row['employee_name']).group(2)
                    print('Name and surname reworked: ', name, ', ', surname)
                
                    ws['B'+str(i)].value = row['date'].strftime('%d/%m')
                    ws['C'+str(i)].value = row['tipo_giorno']
                    ws['D'+str(i)].value = row['ore_lavoro']
                    ws['E'+str(i)].value = row['ore_viaggio']      
                    ws['J'+str(i)].value = 'Vedi rapportini allegati'
                    ws['N'+str(i)].value = name + ' ' + surname
                    print('Values added')
                
                    if row['flag_suspicious'] == True:
                    
                        ws['B'+str(i)].font = Font(bold=True)
                        ws['C'+str(i)].font = Font(bold=True)
                        ws['D'+str(i)].font = Font(bold=True)
                        ws['E'+str(i)].font = Font(bold=True)
                        ws['J'+str(i)].font = Font(bold=True)
                        ws['N'+str(i)].font = Font(bold=True)
                    
                        ws['B'+str(i)].font = Font(color="00FF0000")
                        ws['C'+str(i)].font = Font(color="00FF0000")
                        ws['D'+str(i)].font = Font(color="00FF0000")
                        ws['E'+str(i)].font = Font(color="00FF0000")
                        ws['J'+str(i)].font = Font(color="00FF0000")
                        ws['N'+str(i)].font = Font(color="00FF0000")

                    i = i+1
                
                print(f'Fine {task_category} - {partner} - {project}')
        
            thin = Side(border_style="thin", color="000000")
        
            for i in range(2, 15):
                cell = ws.cell(column=i, row=1)    
                cell.border = Border(bottom = Side(border_style='thick', color='FF000000'))
        
            print('Pre - save the file')          

            name_file_riassunto = str(year) + '-' + str(month) + '_' + task_category.upper() + '_Riassunto.xlsx'
            print('name_file_riassunto: ', name_file_riassunto)

            month_dir = path_output + str(year) + '_' + str(month) + '/'

            wb.remove(wb.worksheets[0])
            print('name_file to save: ', month_dir + name_file_riassunto)
            wb.save(month_dir + name_file_riassunto)
            wb.close()
            del wb

            xlsx_abs = os.path.abspath(month_dir + name_file_riassunto)
            xls_abs  = os.path.abspath(month_dir + name_file_riassunto[:-1])
            # Spire loads the whole workbook into the .NET heap, so Dispose() +
            # gc.collect() after each task_category conversion keeps peak memory
            # flat across the Assistenza/Intervento iterations (Render free tier
            # is 512MB). The riassunto is one multi-sheet file per task_category
            # by design, so it can't be split per partner like the Peve files.
            try:
                print('start of spire.Workbook')
                workbook = spire.Workbook()
                workbook.LoadFromFile(xlsx_abs)
                workbook.SaveToFile(xls_abs, ExcelVersion.Version97to2003)
                workbook.Dispose()
                del workbook
                os.remove(month_dir + name_file_riassunto)
                print(f'Riassunto XLS saved: {xls_abs}')
                try:
                    from storage import upload_file, to_storage_key
                    upload_file(xls_abs, to_storage_key(xls_abs))
                    print(f'Uploaded to Supabase: {to_storage_key(xls_abs)}')
                except Exception as sup_err:
                    print(f'Supabase upload failed (non-fatal): {sup_err}')
            except Exception as spire_err:
                print(f'Spire conversion failed, keeping .xlsx: {spire_err}')
                print(f'Riassunto XLSX saved: {xlsx_abs}')
                try:
                    from storage import upload_file, to_storage_key
                    upload_file(xlsx_abs, to_storage_key(xlsx_abs))
                    print(f'Uploaded to Supabase: {to_storage_key(xlsx_abs)}')
                except Exception as sup_err:
                    print(f'Supabase upload failed (non-fatal): {sup_err}')
            finally:
                gc.collect()

    except Exception as e:
        print(f"Error in create_riassunto: ", e)
        raise




